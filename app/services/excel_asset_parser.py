"""Excel 资产解析服务。"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


EMU_PER_PIXEL = 9525


@dataclass(slots=True)
class ExcelAssetRow:
    """Excel 中的一行资产数据。"""

    sheet_name: str
    row_number: int
    fields: dict[str, str]
    images: list[dict[str, Any]] = field(default_factory=list)


class ExcelAssetParser:
    """把 Excel 文件解析成一行一行的资产数据。"""

    def parse(self, *, file_name: str, content: bytes) -> list[ExcelAssetRow]:
        """解析 Excel 文件，返回资产行列表。"""

        workbook = load_workbook(BytesIO(content), data_only=True)
        asset_rows: list[ExcelAssetRow] = []

        for sheet in workbook.worksheets:
            if self._should_skip_sheet(sheet.title):
                continue

            rows = self._read_sheet(sheet)
            if not rows:
                continue

            header_index = self._find_header_row(rows)
            headers = rows[header_index]
            pending_rows: list[tuple[int, dict[str, str]]] = []

            for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
                if self._is_empty_row(row):
                    continue

                fields = self._build_fields(headers=headers, row=row)
                if fields:
                    pending_rows.append((row_number, fields))

            valid_row_numbers = [row_number for row_number, _ in pending_rows]
            images_by_row = self._collect_images_by_row(
                sheet,
                headers=headers,
                valid_row_numbers=valid_row_numbers,
            )

            for row_number, fields in pending_rows:
                asset_rows.append(
                    ExcelAssetRow(
                        sheet_name=sheet.title,
                        row_number=row_number,
                        fields=fields,
                        images=images_by_row.get(row_number, []),
                    )
                )

        return asset_rows

    # def _read_sheet(self, sheet) -> list[list[str]]:
    #     """读取工作表中的单元格文本。"""

    #     rows: list[list[str]] = []
    #     for row_index in range(1, sheet.max_row + 1):
    #         row_values = [
    #             self._normalize_cell(sheet.cell(row=row_index, column=column_index).value)
    #             for column_index in range(1, sheet.max_column + 1)
    #         ]
    #         rows.append(row_values)
    #     return rows
    def _read_sheet(self, sheet) -> list[list[str]]:
        """读取工作表中的单元格文本。"""

        rows: list[list[str]] = []
        for row_index in range(1, sheet.max_row + 1):
            row_values = [
                self._normalize_cell(
                    self._get_cell_value(
                        sheet,
                        row_index=row_index,
                        column_index=column_index,
                    )
                )
                for column_index in range(1, sheet.max_column + 1)
            ]
            rows.append(row_values)
        return rows
    
    def _get_cell_value(self, sheet, *, row_index: int, column_index: int):
        """读取单元格值；如果是合并单元格，则使用合并区域左上角的值。"""

        cell = sheet.cell(row=row_index, column=column_index)
        if cell.value is not None:
            return cell.value

        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return sheet.cell(
                    row=merged_range.min_row,
                    column=merged_range.min_col,
                ).value

        return None

    def _collect_images_by_row(
        self,
        sheet,
        *,
        headers: list[str],
        valid_row_numbers: list[int],
    ) -> dict[int, list[dict[str, Any]]]:
        """按资产行收集 Excel 图片位置信息。

        openpyxl 的图片 anchor 通常是左上角位置，但图片可能跨行跨列。
        行归属用图片显示区域中心点判断，避免跨行图片挂到上一行。
        列归属用左上角 anchor 判断，避免宽图跨列后被误判成右侧列。
        """

        images_by_row: dict[int, list[dict[str, Any]]] = {}

        for image_index, image in enumerate(getattr(sheet, "_images", [])):
            anchor = getattr(image, "anchor", None)
            marker = getattr(anchor, "_from", None)
            if marker is None:
                continue

            left, top = self._marker_to_pixels(sheet, marker)
            width, height = self._image_display_size_pixels(sheet, image)
            center_y = top + height / 2

            assigned_row_number = self._find_nearest_asset_row(
                sheet,
                center_y=center_y,
                valid_row_numbers=valid_row_numbers,
            )
            assigned_column_number = marker.col + 1
            column_header = ""
            if 0 <= assigned_column_number - 1 < len(headers):
                column_header = headers[assigned_column_number - 1]

            image_info = {
                "sheet_name": sheet.title,
                "image_index": image_index,
                "row": assigned_row_number,
                "col": assigned_column_number,
                "column_header": column_header,
                "format": getattr(image, "format", None),
                "width": getattr(image, "width", None),
                "height": getattr(image, "height", None),
            }
            images_by_row.setdefault(assigned_row_number, []).append(image_info)

        return images_by_row

    def _find_header_row(self, rows: list[list[str]]) -> int:
        """根据行结构猜测表头行。"""

        best_index = 0
        best_score = -1

        for index, row in enumerate(rows[:30]):
            cells = [cell for cell in row if cell]
            if not cells:
                continue

            short_count = sum(1 for cell in cells if len(cell) <= 30)
            long_count = sum(1 for cell in cells if len(cell) > 80)
            unique_count = len(set(cells))
            duplicate_count = len(cells) - unique_count

            score = 0
            score += len(cells) * 2
            score += short_count
            score -= long_count * 4
            score -= duplicate_count * 5
            if len(cells) <= 2:
                score -= 10

            if score > best_score:
                best_score = score
                best_index = index

        return best_index

    @staticmethod
    def _build_fields(*, headers: list[str], row: list[str]) -> dict[str, str]:
        """把一行数据转换成 表头 -> 单元格值。"""

        fields: dict[str, str] = {}
        for index, header in enumerate(headers):
            value = row[index] if index < len(row) else ""
            if header and value:
                fields[header] = value
        return fields

    def _marker_to_pixels(self, sheet, marker) -> tuple[float, float]:
        """把 openpyxl 的 marker 转换成近似像素坐标。"""

        left = sum(self._column_width_pixels(sheet, column) for column in range(1, marker.col + 1))
        top = sum(self._row_height_pixels(sheet, row) for row in range(1, marker.row + 1))
        left += getattr(marker, "colOff", 0) / EMU_PER_PIXEL
        top += getattr(marker, "rowOff", 0) / EMU_PER_PIXEL
        return left, top

    def _image_display_size_pixels(self, sheet, image) -> tuple[float, float]:
        """获取图片在 Excel 中的显示尺寸。"""

        anchor = getattr(image, "anchor", None)
        ext = getattr(anchor, "ext", None)
        if ext is not None:
            return ext.cx / EMU_PER_PIXEL, ext.cy / EMU_PER_PIXEL

        from_marker = getattr(anchor, "_from", None)
        to_marker = getattr(anchor, "_to", None)
        if from_marker is not None and to_marker is not None:
            left, top = self._marker_to_pixels(sheet, from_marker)
            right, bottom = self._marker_to_pixels(sheet, to_marker)
            return max(right - left, 0), max(bottom - top, 0)

        return float(getattr(image, "width", 0) or 0), float(getattr(image, "height", 0) or 0)

    def _find_nearest_asset_row(
        self,
        sheet,
        *,
        center_y: float,
        valid_row_numbers: list[int],
    ) -> int:
        """根据图片中心点找到对应资产行。"""

        if not valid_row_numbers:
            return 1

        best_row = valid_row_numbers[0]
        best_distance = float("inf")

        for row_number in valid_row_numbers:
            top = self._row_top_pixels(sheet, row_number)
            bottom = top + self._row_height_pixels(sheet, row_number)
            if top <= center_y <= bottom:
                return row_number

            row_center = (top + bottom) / 2
            distance = abs(center_y - row_center)
            if distance < best_distance:
                best_distance = distance
                best_row = row_number

        return best_row

    def _row_top_pixels(self, sheet, row_number: int) -> float:
        """计算某一行顶部的近似像素坐标。"""

        return sum(self._row_height_pixels(sheet, row) for row in range(1, row_number))

    @staticmethod
    def _row_height_pixels(sheet, row_number: int) -> float:
        """把 Excel 行高转换成近似像素。"""

        height_points = sheet.row_dimensions[row_number].height
        if height_points is None:
            height_points = sheet.sheet_format.defaultRowHeight or 15
        return float(height_points) * 96 / 72

    @staticmethod
    def _column_width_pixels(sheet, column_number: int) -> float:
        """把 Excel 列宽转换成近似像素。"""

        column_letter = get_column_letter(column_number)
        width = sheet.column_dimensions[column_letter].width
        if width is None:
            width = sheet.sheet_format.defaultColWidth or 8.43
        return float(width) * 7 + 5

    @staticmethod
    def _normalize_cell(value: Any) -> str:
        """把单元格值转换成字符串。"""

        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _is_empty_row(row: list[str]) -> bool:
        """判断是否为空行。"""

        return not any(cell.strip() for cell in row)

    @staticmethod
    def _should_skip_sheet(sheet_name: str) -> bool:
        """跳过说明类工作表。"""

        skip_keywords = ["使用说明", "说明", "README", "readme"]
        return any(keyword in sheet_name for keyword in skip_keywords)
