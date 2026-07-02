"""Excel 内嵌图片提取服务。"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


@dataclass(slots=True)
class ExcelImageFile:
    """从 Excel 中提取出的一张图片。"""

    sheet_name: str
    image_index: int
    content: bytes
    extension: str
    content_type: str


class ExcelImageExtractor:
    """根据 sheet_name + image_index 从 Excel 中提取图片 bytes。"""

    def extract_one(
        self,
        *,
        excel_path: str,
        sheet_name: str,
        image_index: int,
    ) -> ExcelImageFile:
        """提取一张 Excel 内嵌图片。"""

        path = Path(excel_path)
        workbook = load_workbook(path, data_only=True)

        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Excel sheet not found: {sheet_name}")

        sheet = workbook[sheet_name]
        images = list(getattr(sheet, "_images", []))

        if image_index < 0 or image_index >= len(images):
            raise IndexError(
                f"image_index out of range: {image_index}, "
                f"sheet={sheet_name}, image_count={len(images)}"
            )

        image = images[image_index]
        content = image._data()
        extension = self._normalize_extension(getattr(image, "format", None))

        return ExcelImageFile(
            sheet_name=sheet_name,
            image_index=image_index,
            content=content,
            extension=extension,
            content_type=self._content_type(extension),
        )

    @staticmethod
    def _normalize_extension(value: str | None) -> str:
        extension = (value or "png").lower().strip(".")
        if extension == "jpeg":
            return "jpg"
        return extension

    @staticmethod
    def _content_type(extension: str) -> str:
        mapping = {
            "jpg": "image/jpeg",
            "jpeg": "image/jpeg",
            "png": "image/png",
            "gif": "image/gif",
            "webp": "image/webp",
        }
        return mapping.get(extension, "application/octet-stream")
