"""分镜 Excel 解析器。

职责：
1. 打开 Excel 文件。
2. 遍历每个 sheet。
3. 让模型识别表头。
4. 把每一行转换成标准 StoryboardExcelRow。

注意：
这里暂时不负责数据库入库、不负责向量化、不负责图片上传。
这样代码分层会清楚很多。
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.services.Mapping.storyboard_field_mapper import (
    StoryboardExcelRow,
    StoryboardFieldMapper,
)
from app.services.Mapping.storyboard_header_mapper import StoryboardHeaderMapper


class StoryboardExcelParser:
    """解析 Excel 中的分镜数据。"""

    def __init__(
        self,
        header_mapper: StoryboardHeaderMapper,
        field_mapper: StoryboardFieldMapper,
    ) -> None:
        # 表头识别交给独立类，后面可以单独替换/优化模型提示词。
        self.header_mapper = header_mapper

        # 行字段映射交给独立类，避免 parser 里堆太多业务字段处理。
        self.field_mapper = field_mapper

    def parse(self, excel_path: Path) -> list[StoryboardExcelRow]:
        """解析整个 Excel，返回标准化后的分镜行列表。"""

        workbook = load_workbook(excel_path, data_only=True)

        rows: list[StoryboardExcelRow] = []

        for sheet in workbook.worksheets:
            rows.extend(self._parse_sheet(sheet))

        return rows

    def _parse_sheet(self, sheet: Worksheet) -> list[StoryboardExcelRow]:
        """解析单个 sheet。"""

        # 第一步先取第一行作为候选表头。
        # 后面如果你的 Excel 表头不一定在第一行，我们再增强成自动找表头。
        header_result = self._find_header_row(sheet)

        # 找不到分镜表头，说明这个 sheet 不是分镜表，或者格式暂时不支持。
        if header_result is None:
            return []

        header_row, headers = header_result
        
        if not self._is_storyboard_sheet(sheet.title, headers):
            return []

        # 让模型判断这些原始表头分别对应哪些分镜标准字段。
        header_map = self.header_mapper.map_headers(
            sheet_name=sheet.title,
            headers=headers,
        )

        parsed_rows: list[StoryboardExcelRow] = []

        # 从第二行开始读数据，因为第一行先当作表头。
        for row_number in range(header_row + 1, sheet.max_row + 1):
            raw_fields = self._read_row_fields(
                sheet=sheet,
                headers=headers,
                row_number=row_number,
            )

            # 如果这一行全空，就跳过。
            if not any(raw_fields.values()):
                continue

            parsed_rows.append(
                self.field_mapper.map_fields(
                    sheet_name=sheet.title,
                    row_number=row_number,
                    fields=raw_fields,
                    header_map=header_map,
                )
            )

        return parsed_rows

    def _read_header_row(self, sheet: Worksheet, header_row: int) -> list[str]:
        """读取表头行。"""

        headers: list[str] = []

        for cell in sheet[header_row]:
            value = self._normalize_cell_value(cell.value)
            if value:
                headers.append(value)

        return headers
    

    def _find_header_row(self, sheet: Worksheet) -> tuple[int, list[str]] | None:
        """在 sheet 前几行里自动寻找分镜表头行。

        真实 Excel 里经常会有：
        - 第 1 行是标题
        - 第 2 行是说明
        - 第 3 行才是真正表头

        所以这里不固定只读第 1 行，而是在前 10 行里找最像表头的行。
        """

        max_scan_rows = min(sheet.max_row, 10)

        for row_number in range(1, max_scan_rows + 1):
            headers = self._read_header_row(sheet, header_row=row_number)

            if not headers:
                continue

            if self._is_storyboard_sheet(sheet.title, headers):
                return row_number, headers

        return None

    def _read_row_fields(
        self,
        *,
        sheet: Worksheet,
        headers: list[str],
        row_number: int,
    ) -> dict[str, str]:
        """按表头读取某一行的数据。"""

        fields: dict[str, str] = {}

        for column_index, header in enumerate(headers, start=1):
            value = sheet.cell(row=row_number, column=column_index).value
            fields[header] = self._normalize_cell_value(value)

        return fields

    @staticmethod
    def _normalize_cell_value(value: object) -> str:
        """统一 Excel 单元格值，空值转成空字符串。"""

        return str(value or "").strip()
    

    def _is_storyboard_sheet(self, sheet_name: str, headers: list[str]) -> bool:
        """判断当前 sheet 是否像分镜表。

        这里先用轻量规则过滤：
        1. sheet 名包含“分镜”，基本可以认为是分镜表。
        2. 或者表头里同时出现几个分镜强特征字段，比如“镜号/画面/镜头/台词”。

        注意：
        这一步不是做字段映射，只是避免把人物表、场景表也送进分镜解析。
        真正的字段含义仍然交给 StoryboardHeaderMapper 用模型识别。
        """

        sheet_name_text = sheet_name.strip()

        if "分镜" in sheet_name_text:
            return True

        header_text = " ".join(headers)

        storyboard_keywords = [
            "镜号",
            "镜头",
            "画面",
            "台词",
            "景别",
            "视角",
            "音效",
        ]

        matched_count = sum(
            1 for keyword in storyboard_keywords if keyword in header_text
        )

        # 命中 2 个以上，才认为它像分镜表。
        # 这样可以避免“场景表”里偶然有一个画面字段就误判。
        return matched_count >= 3
    

