"""评测 04_query_answer+confusion.xlsx 的图片语义向量召回效果。

这个脚本适配当前 Excel 的结构：
1. 第 2-21 行是 20 条 query，同时这些行也嵌入了标准答案图片。
2. 第 22-121 行是 100 张混淆候选图片，行 ID 是 CONF001-CONF100。
3. 候选池一共 120 张图片：20 张标准答案图片 + 100 张混淆图片。

评测流程：
1. 用视觉模型理解每张候选图片，生成中文语义描述。
2. 把图片语义描述向量化，作为候选召回块。
3. 把每条 query 原文向量化。
4. 用 query 向量和所有候选图片描述向量计算余弦相似度。
5. 输出每条 query 的 Top5 召回块、分数、正确答案排名，并回填 Excel。

注意：脚本只做离线评测，不写数据库，也不写 Milvus。
"""

from __future__ import annotations

import argparse
import base64
import hashlib
import json
import math
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import httpx
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.services.vector.doubao_embedding_service import DoubaoEmbeddingService


DOCS_DIR = Path("docs") / "rag提升文档"
DEFAULT_EXCEL_FILE = Path.home() / "Desktop" / "04_query_answer+confusion.xlsx"
DEFAULT_OUTPUT_EXCEL_FILE = (
    Path.home() / "Desktop" / "04_query_answer+confusion_image_semantic_result.xlsx"
)
DEFAULT_CACHE_FILE = DOCS_DIR / "04_query_answer_confusion_image_semantic_cache.json"
DEFAULT_DATASET_FILE = DOCS_DIR / "04_query_answer_confusion_image_semantic_dataset.json"
DEFAULT_RESULT_JSON_FILE = DOCS_DIR / "04_query_answer_confusion_image_semantic_eval_result.json"
DEFAULT_RESULT_MD_FILE = DOCS_DIR / "04_query_answer_confusion_image_semantic_eval_result.md"


@dataclass(slots=True)
class QueryCase:
    """一条待评测的问题。"""

    excel_row: int
    query_id: str
    query_text: str
    expected_answer_asset_id: str
    asset_type: str


@dataclass(slots=True)
class ImageCandidate:
    """一张候选图片，以及它被视觉模型理解后的向量文本。"""

    excel_row: int
    candidate_id: str
    candidate_group: str
    asset_type: str
    mapped_answer_asset_id: str | None
    visual_summary: str
    image_search_text: str
    vector: list[float]


def clean_text(value: Any) -> str:
    """把 Excel 单元格、模型输出等值统一成单行文本，避免换行影响写表。"""

    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\r", " ").replace("\n", " ")).strip()


def is_null_text(value: str) -> bool:
    """判断 Excel 里的空字符串/null 文本。"""

    return clean_text(value).lower() in {"", "null", "none", "nan"}


def read_env(path: Path) -> dict[str, str]:
    """读取 .env 文件；这里只读配置，不打印密钥。"""

    data: dict[str, str] = {}
    if not path.exists():
        return data

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        data[key.strip()] = value.strip().strip('"').strip("'")
    return data


def load_cache(path: Path) -> dict[str, Any]:
    """读取缓存；没有缓存时初始化两个空间：vision 和 embedding。"""

    if not path.exists():
        return {"vision": {}, "embedding": {}}

    data = json.loads(path.read_text(encoding="utf-8"))
    data.setdefault("vision", {})
    data.setdefault("embedding", {})
    return data


def save_cache(path: Path, cache: dict[str, Any]) -> None:
    """保存缓存。每处理完一张图就保存一次，方便中断后继续跑。"""

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(cache, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def header_index(worksheet: Any) -> dict[str, int]:
    """读取表头位置，返回 header -> column_index。"""

    return {clean_text(cell.value): index for index, cell in enumerate(worksheet[1], start=1)}


def parse_cases_and_candidates(
    workbook: openpyxl.Workbook,
) -> tuple[Any, list[QueryCase], list[dict[str, Any]], dict[int, bytes]]:
    """从 Excel 解析 query 行、候选图片行和图片二进制。

    当前表的 Q 行既是问题，也是标准答案图片行：
    Q001 行图片 -> ANS001
    Q002 行图片 -> ANS002
    ...
    CONF001-CONF100 行只参与召回排序，不作为正确答案。
    """

    worksheet = workbook["test"] if "test" in workbook.sheetnames else workbook.active
    columns = header_index(worksheet)
    required = {"query_id", "query", "answer_asset_id", "asset_type"}
    missing = sorted(required.difference(columns))
    if missing:
        raise ValueError(f"Excel 缺少必要列：{', '.join(missing)}")

    image_bytes_by_row: dict[int, bytes] = {}
    for image in getattr(worksheet, "_images", []):
        row_number = image.anchor._from.row + 1
        image_bytes_by_row[row_number] = image._data()

    cases: list[QueryCase] = []
    candidate_rows: list[dict[str, Any]] = []

    for row_number in range(2, worksheet.max_row + 1):
        row_id = clean_text(worksheet.cell(row_number, columns["query_id"]).value)
        query_text = clean_text(worksheet.cell(row_number, columns["query"]).value)
        answer_asset_id = clean_text(worksheet.cell(row_number, columns["answer_asset_id"]).value)
        asset_type = clean_text(worksheet.cell(row_number, columns["asset_type"]).value)

        if not row_id:
            continue

        is_answer_row = row_id.startswith("Q") and not is_null_text(answer_asset_id)

        # Q 开头的是评测问题行，同时它自己的图片也是标准答案候选。
        if is_answer_row:
            cases.append(
                QueryCase(
                    excel_row=row_number,
                    query_id=row_id,
                    query_text=query_text,
                    expected_answer_asset_id=answer_asset_id,
                    asset_type=asset_type,
                )
            )

        # 只要这一行带图，就加入候选池。Q 行带标准答案 ID，CONF 行没有标准答案 ID。
        if row_number in image_bytes_by_row:
            mapped_answer_asset_id = answer_asset_id if is_answer_row else None
            candidate_group = "标准答案图片" if is_answer_row else query_text

            candidate_rows.append(
                {
                    "excel_row": row_number,
                    "candidate_id": row_id,
                    "candidate_group": candidate_group,
                    "asset_type": asset_type,
                    "mapped_answer_asset_id": mapped_answer_asset_id,
                }
            )

    if not cases:
        raise ValueError("没有解析到 Q 开头的问题行。")
    if not candidate_rows:
        raise ValueError("没有解析到带图片的候选行。")

    return worksheet, cases, candidate_rows, image_bytes_by_row


def vision_prompt(candidate_row: dict[str, Any]) -> str:
    """构造视觉理解提示词。

    不把 mapped_answer_asset_id 放进提示词，避免模型描述里混入答案标签。
    """

    return f"""
你是中文短剧素材库的图片标注员。请只根据图片内容生成中文语义描述，不要编造图片里看不到的信息。

已知素材类型：{candidate_row["asset_type"] or "未知"}

请输出严格 JSON，不要 Markdown，不要解释：
{{
  "visual_summary": "一句话描述图片主体、风格、关键视觉特征和可检索用途",
  "subject": "主体是什么，例如都市男主、古风女主、医院病房、修炼仪式场景、道具等",
  "asset_type": "角色/角色变体/场景/道具/未知",
  "gender": "男/女/未知/不适用",
  "age_range": "年龄段或不适用",
  "hair": "发型、发色、头部特征",
  "clothing": "服装、鞋帽、配饰",
  "colors": ["主色1", "主色2"],
  "scene_space": "场景空间或背景",
  "style": ["都市短剧", "仙侠", "古风", "恐怖", "现实", "悬疑"],
  "mood": "气质氛围，例如清冷、压迫、甜美、诡异、破败",
  "props": ["可见道具"],
  "search_tags": ["适合检索的中文标签"],
  "search_text": "把以上信息合并成一段适合向量检索的中文描述，重点包含主体、服装、场景、风格、气质、用途"
}}
""".strip()


def parse_json_object(text: str) -> dict[str, Any]:
    """尽量从模型输出里解析 JSON；失败时把原文当描述兜底。"""

    content = text.strip()
    if content.startswith("```"):
        content = re.sub(r"^```(?:json)?", "", content).strip()
        content = re.sub(r"```$", "", content).strip()

    try:
        parsed = json.loads(content)
        if isinstance(parsed, dict):
            return parsed
    except json.JSONDecodeError:
        pass

    match = re.search(r"\{.*\}", content, flags=re.S)
    if match:
        try:
            parsed = json.loads(match.group(0))
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            pass

    return {
        "visual_summary": clean_text(text),
        "search_text": clean_text(text),
        "search_tags": [],
    }


def describe_image(
    *,
    candidate_row: dict[str, Any],
    image_bytes: bytes,
    cache: dict[str, Any],
    model: str,
    api_key: str,
    base_url: str,
    timeout: int,
) -> dict[str, Any]:
    """调用视觉模型生成图片描述，并写入缓存。"""

    image_hash = hashlib.sha256(image_bytes).hexdigest()
    cache_key = f"{candidate_row['candidate_id']}:{image_hash}"
    if cache_key in cache["vision"]:
        return cache["vision"][cache_key]

    image_b64 = base64.b64encode(image_bytes).decode("ascii")
    payload = {
        "model": model,
        "messages": [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": vision_prompt(candidate_row)},
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:image/jpeg;base64,{image_b64}"},
                    },
                ],
            }
        ],
        "temperature": 0,
    }

    response = httpx.post(
        f"{base_url.rstrip('/')}/chat/completions",
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=timeout,
    )
    if response.status_code >= 400:
        raise RuntimeError(
            f"Vision model failed for {candidate_row['candidate_id']}: "
            f"status={response.status_code}, body={response.text}"
        )

    data = response.json()
    content = data["choices"][0]["message"]["content"]
    parsed = parse_json_object(content)
    parsed["raw_content"] = content
    parsed["model"] = data.get("model", model)
    cache["vision"][cache_key] = parsed
    return parsed


def embedding_for_text(
    *,
    service: DoubaoEmbeddingService,
    cache: dict[str, Any],
    prefix: str,
    text: str,
) -> list[float]:
    """向量化文本，并缓存。

    prefix 用来区分 query 和 image_search_text，避免同一段文本在不同用途下互相覆盖。
    """

    key = f"{prefix}:{hashlib.sha256(text.encode('utf-8')).hexdigest()}"
    if key not in cache["embedding"]:
        cache["embedding"][key] = service.embed_text(text)
    return [float(item) for item in cache["embedding"][key]]


def cosine_similarity(left: list[float], right: list[float]) -> float:
    """计算两个向量的余弦相似度，值越大代表语义越接近。"""

    dot = sum(a * b for a, b in zip(left, right))
    left_norm = math.sqrt(sum(a * a for a in left))
    right_norm = math.sqrt(sum(b * b for b in right))
    if left_norm == 0 or right_norm == 0:
        return 0.0
    return dot / (left_norm * right_norm)


def candidate_display(candidate: dict[str, Any]) -> str:
    """生成 Excel 里可读的候选块名称。"""

    mapped = candidate.get("mapped_answer_asset_id") or "干扰项"
    asset_type = candidate.get("asset_type") or "未知"
    return f"{candidate['candidate_id']}/{mapped}/{asset_type}"


def build_dataset_and_evaluate(
    *,
    excel_file: Path,
    cache_file: Path,
    dataset_file: Path,
    result_json_file: Path,
    result_md_file: Path,
    output_excel_file: Path | None,
    top_k: int,
    vision_model: str | None,
    timeout: int,
) -> dict[str, Any]:
    """执行完整的图片语义检索评测。"""

    env = read_env(PROJECT_ROOT / ".env")
    api_key = env.get("ARK_API_KEY")
    if not api_key:
        raise ValueError("Missing ARK_API_KEY in .env")

    # 兼容你之前写过的 DOUBAO_LLM_MODE_2.0lite 变量名。
    model = (
        vision_model
        or env.get("DOUBAO_LLM_MODEL_2_0_LITE")
        or env.get("DOUBAO_LLM_MODE_2.0lite")
        or env.get("DOUBAO_VISION_MODEL")
        or env.get("DOUBAO_LLM_MODEL")
    )
    if not model:
        raise ValueError("Missing vision model endpoint in .env")

    base_url = env.get("ARK_BASE_URL", "https://ark.cn-beijing.volces.com/api/v3")
    workbook = openpyxl.load_workbook(excel_file)
    _worksheet, cases, candidate_rows, image_bytes_by_row = parse_cases_and_candidates(workbook)

    cache = load_cache(cache_file)
    embedding_service = DoubaoEmbeddingService()

    candidates: list[ImageCandidate] = []
    dataset_candidates: list[dict[str, Any]] = []

    for index, row in enumerate(candidate_rows, start=1):
        print(f"[vision] {index}/{len(candidate_rows)} {row['candidate_id']}")
        image_bytes = image_bytes_by_row[row["excel_row"]]
        visual = describe_image(
            candidate_row=row,
            image_bytes=image_bytes,
            cache=cache,
            model=model,
            api_key=api_key,
            base_url=base_url,
            timeout=timeout,
        )
        visual_summary = clean_text(visual.get("visual_summary"))
        image_search_text = clean_text(visual.get("search_text")) or visual_summary
        if not image_search_text:
            image_search_text = clean_text(visual.get("raw_content"))

        vector = embedding_for_text(
            service=embedding_service,
            cache=cache,
            prefix="image_search_text",
            text=image_search_text,
        )
        save_cache(cache_file, cache)

        candidate = ImageCandidate(
            excel_row=row["excel_row"],
            candidate_id=row["candidate_id"],
            candidate_group=row["candidate_group"],
            asset_type=row["asset_type"],
            mapped_answer_asset_id=row["mapped_answer_asset_id"],
            visual_summary=visual_summary,
            image_search_text=image_search_text,
            vector=vector,
        )
        candidates.append(candidate)
        dataset_candidates.append(
            {
                **row,
                "visual": visual,
                "visual_summary": visual_summary,
                "image_search_text": image_search_text,
            }
        )

    eval_items: list[dict[str, Any]] = []
    for index, case in enumerate(cases, start=1):
        print(f"[query] {index}/{len(cases)} {case.query_id}")
        query_vector = embedding_for_text(
            service=embedding_service,
            cache=cache,
            prefix="query",
            text=case.query_text,
        )
        save_cache(cache_file, cache)

        ranked = sorted(
            (
                {
                    "excel_row": candidate.excel_row,
                    "candidate_id": candidate.candidate_id,
                    "candidate_group": candidate.candidate_group,
                    "asset_type": candidate.asset_type,
                    "mapped_answer_asset_id": candidate.mapped_answer_asset_id,
                    "visual_summary": candidate.visual_summary,
                    "image_search_text": candidate.image_search_text,
                    "score": cosine_similarity(query_vector, candidate.vector),
                }
                for candidate in candidates
            ),
            key=lambda item: item["score"],
            reverse=True,
        )

        hit_rank = None
        hit_score = None
        for rank, candidate in enumerate(ranked, start=1):
            if candidate["mapped_answer_asset_id"] == case.expected_answer_asset_id:
                hit_rank = rank
                hit_score = candidate["score"]
                break

        eval_items.append(
            {
                "excel_row": case.excel_row,
                "query_id": case.query_id,
                "query_text": case.query_text,
                "expected_answer_asset_id": case.expected_answer_asset_id,
                "asset_type": case.asset_type,
                "hit_rank": hit_rank,
                "hit_score": hit_score,
                "hit_at_top_k": hit_rank is not None and hit_rank <= top_k,
                "top_k": ranked[:top_k],
            }
        )

    total = len(eval_items)
    hit_count = sum(1 for item in eval_items if item["hit_at_top_k"])
    hit_at_1 = sum(1 for item in eval_items if item["hit_rank"] == 1)
    hit_at_3 = sum(
        1 for item in eval_items if item["hit_rank"] is not None and item["hit_rank"] <= 3
    )
    mrr = sum(
        0.0 if item["hit_rank"] is None else 1.0 / item["hit_rank"]
        for item in eval_items
    ) / total

    dataset = {
        "dataset_name": excel_file.stem,
        "source_excel": str(excel_file),
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "vision_model": model,
        "answer_mapping_rule": "Q001-Q020 行图片映射各自行内 answer_asset_id；CONF001-CONF100 为混淆候选。",
        "vectorization_rule": {
            "query_text": "使用 query 字段原文向量化。",
            "candidate_text": "使用视觉模型生成的 image_search_text 向量化。",
            "hit_rule": "按 mapped_answer_asset_id 是否等于 query 行的 answer_asset_id 判断命中。",
        },
        "queries": [
            {
                "excel_row": case.excel_row,
                "query_id": case.query_id,
                "query_text": case.query_text,
                "expected_answer_asset_id": case.expected_answer_asset_id,
                "asset_type": case.asset_type,
            }
            for case in cases
        ],
        "candidates": dataset_candidates,
    }
    result = {
        "dataset_name": excel_file.stem,
        "evaluated_at": datetime.now().isoformat(timespec="seconds"),
        "vision_model": model,
        "top_k": top_k,
        "source_excel": str(excel_file),
        "output_excel": "" if output_excel_file is None else str(output_excel_file),
        "answer_mapping_rule": dataset["answer_mapping_rule"],
        "metrics": {
            "total": total,
            "hit_count": hit_count,
            "hit_at_k": hit_count / total if total else 0.0,
            "hit_at_1": hit_at_1 / total if total else 0.0,
            "hit_at_3": hit_at_3 / total if total else 0.0,
            "mrr": mrr,
        },
        "items": eval_items,
        "candidates": dataset_candidates,
    }

    dataset_file.parent.mkdir(parents=True, exist_ok=True)
    dataset_file.write_text(
        json.dumps(dataset, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    result_json_file.parent.mkdir(parents=True, exist_ok=True)
    result_json_file.write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    write_markdown_result(result=result, path=result_md_file)

    if output_excel_file is not None:
        backfill_excel(
            excel_file=excel_file,
            output_excel_file=output_excel_file,
            result=result,
        )

    return result


def write_markdown_result(*, result: dict[str, Any], path: Path) -> None:
    """写一份方便直接看的 Markdown 评测报告。"""

    metrics = result["metrics"]
    lines = [
        "# 04 图片语义描述向量检索评测结果",
        "",
        f"- 数据集：`{result['dataset_name']}`",
        f"- 视觉模型：`{result['vision_model']}`",
        f"- TopK：`{result['top_k']}`",
        f"- 评测时间：{result['evaluated_at']}",
        f"- 答案映射规则：{result['answer_mapping_rule']}",
        "",
        "## 指标",
        "",
        "| 指标 | 数值 |",
        "|---|---:|",
        f"| total | {metrics['total']} |",
        f"| hit_count | {metrics['hit_count']} |",
        f"| Hit@{result['top_k']} | {metrics['hit_at_k']:.2%} |",
        f"| Hit@1 | {metrics['hit_at_1']:.2%} |",
        f"| Hit@3 | {metrics['hit_at_3']:.2%} |",
        f"| MRR | {metrics['mrr']:.4f} |",
        "",
        "## 明细",
        "",
        "| Query ID | 标准答案 | 命中排名 | 命中分数 | Top1 | Top1 分数 | Top5 |",
        "|---|---|---:|---:|---|---:|---|",
    ]

    for item in result["items"]:
        top_items = item["top_k"]
        top1 = top_items[0] if top_items else {}
        top5 = "；".join(
            f"{rank}.{candidate_display(candidate)}({candidate['score']:.4f})"
            for rank, candidate in enumerate(top_items, start=1)
        )
        lines.append(
            "| {query_id} | {expected} | {hit_rank} | {hit_score} | {top1} | {top1_score} | {top5} |".format(
                query_id=item["query_id"],
                expected=item["expected_answer_asset_id"],
                hit_rank="" if item["hit_rank"] is None else item["hit_rank"],
                hit_score="" if item["hit_score"] is None else f"{item['hit_score']:.6f}",
                top1=candidate_display(top1) if top1 else "",
                top1_score="" if "score" not in top1 else f"{top1['score']:.6f}",
                top5=top5,
            )
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def ensure_headers(worksheet: Any, headers: list[str]) -> dict[str, int]:
    """确保工作表存在回填列，返回 header -> column_index。"""

    existing = header_index(worksheet)
    for header in headers:
        if header not in existing:
            column = worksheet.max_column + 1
            worksheet.cell(row=1, column=column, value=header)
            existing[header] = column
    return existing


def style_header(worksheet: Any) -> None:
    """设置表头样式，让回填结果更好看。"""

    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def adjust_width(worksheet: Any, max_width: int = 70) -> None:
    """简单调整列宽，并开启自动换行。"""

    for column_cells in worksheet.columns:
        letter = column_cells[0].column_letter
        width = max(len(clean_text(cell.value)) for cell in column_cells) + 2
        worksheet.column_dimensions[letter].width = min(max(width, 10), max_width)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def backfill_excel(
    *,
    excel_file: Path,
    output_excel_file: Path,
    result: dict[str, Any],
) -> None:
    """把评测结果写回 Excel。

    为了避免原文件被 Excel 打开导致保存失败，默认写到一个新的结果文件。
    """

    output_excel_file.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook["test"] if "test" in workbook.sheetnames else workbook.active

    result_by_query_row = {item["excel_row"]: item for item in result["items"]}
    candidate_by_row = {item["excel_row"]: item for item in result["candidates"]}

    headers = [
        "图片语义命中排名",
        "图片语义命中分数",
        "图片语义是否Top5命中",
        "图片语义Top1召回块",
        "图片语义Top1分数",
        "图片语义Top5召回块",
        "图片语义Top5分数",
        "图片语义Top5明细",
        "候选映射答案ID",
        "候选图片语义描述",
        "候选图片向量文本",
        "评测规则",
    ]
    columns = ensure_headers(worksheet, headers)

    for row_number in range(2, worksheet.max_row + 1):
        item = result_by_query_row.get(row_number)
        if item:
            top_k = item["top_k"]
            top1 = top_k[0] if top_k else {}
            top5_blocks = "；".join(
                f"{rank}.{candidate_display(candidate)}"
                for rank, candidate in enumerate(top_k, start=1)
            )
            top5_scores = "；".join(
                f"{rank}.{candidate['score']:.6f}"
                for rank, candidate in enumerate(top_k, start=1)
            )
            top5_detail = "；".join(
                (
                    f"{rank}.{candidate_display(candidate)} "
                    f"score={candidate['score']:.6f} "
                    f"summary={candidate['visual_summary']}"
                )
                for rank, candidate in enumerate(top_k, start=1)
            )

            worksheet.cell(row_number, columns["图片语义命中排名"], item["hit_rank"])
            worksheet.cell(row_number, columns["图片语义命中分数"], item["hit_score"])
            worksheet.cell(
                row_number,
                columns["图片语义是否Top5命中"],
                "是" if item["hit_at_top_k"] else "否",
            )
            worksheet.cell(
                row_number,
                columns["图片语义Top1召回块"],
                candidate_display(top1) if top1 else "",
            )
            worksheet.cell(row_number, columns["图片语义Top1分数"], top1.get("score"))
            worksheet.cell(row_number, columns["图片语义Top5召回块"], top5_blocks)
            worksheet.cell(row_number, columns["图片语义Top5分数"], top5_scores)
            worksheet.cell(row_number, columns["图片语义Top5明细"], top5_detail)
            worksheet.cell(row_number, columns["评测规则"], result["answer_mapping_rule"])

        candidate = candidate_by_row.get(row_number)
        if candidate:
            worksheet.cell(
                row_number,
                columns["候选映射答案ID"],
                candidate.get("mapped_answer_asset_id") or "",
            )
            worksheet.cell(
                row_number,
                columns["候选图片语义描述"],
                candidate.get("visual_summary") or "",
            )
            worksheet.cell(
                row_number,
                columns["候选图片向量文本"],
                candidate.get("image_search_text") or "",
            )

    for sheet_name in [
        "image_semantic_summary",
        "image_semantic_top5",
        "image_semantic_candidates",
    ]:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    summary_sheet = workbook.create_sheet("image_semantic_summary")
    summary_sheet.append(["指标", "数值"])
    metrics = result["metrics"]
    for metric, value in [
        ("total", metrics["total"]),
        ("hit_count", metrics["hit_count"]),
        (f"Hit@{result['top_k']}", metrics["hit_at_k"]),
        ("Hit@1", metrics["hit_at_1"]),
        ("Hit@3", metrics["hit_at_3"]),
        ("MRR", metrics["mrr"]),
        ("vision_model", result["vision_model"]),
        ("answer_mapping_rule", result["answer_mapping_rule"]),
    ]:
        summary_sheet.append([metric, value])

    detail_sheet = workbook.create_sheet("image_semantic_top5")
    detail_sheet.append(
        [
            "query_id",
            "query_text",
            "expected_answer_asset_id",
            "hit_rank",
            "hit_score",
            "hit_at_top5",
            "recall_rank",
            "recall_candidate_id",
            "recall_mapped_answer_asset_id",
            "recall_asset_type",
            "recall_score",
            "is_expected",
            "recall_visual_summary",
            "recall_image_search_text",
        ]
    )
    for item in result["items"]:
        expected_id = item["expected_answer_asset_id"]
        for rank, candidate in enumerate(item["top_k"], start=1):
            candidate_answer_id = candidate.get("mapped_answer_asset_id")
            detail_sheet.append(
                [
                    item["query_id"],
                    item["query_text"],
                    expected_id,
                    item["hit_rank"],
                    item["hit_score"],
                    "是" if item["hit_at_top_k"] else "否",
                    rank,
                    candidate["candidate_id"],
                    candidate_answer_id or "",
                    candidate["asset_type"],
                    candidate["score"],
                    "是" if candidate_answer_id == expected_id else "否",
                    candidate["visual_summary"],
                    candidate["image_search_text"],
                ]
            )

    candidate_sheet = workbook.create_sheet("image_semantic_candidates")
    candidate_sheet.append(
        [
            "excel_row",
            "candidate_id",
            "candidate_group",
            "asset_type",
            "mapped_answer_asset_id",
            "visual_summary",
            "image_search_text",
        ]
    )
    for candidate in result["candidates"]:
        candidate_sheet.append(
            [
                candidate["excel_row"],
                candidate["candidate_id"],
                candidate["candidate_group"],
                candidate["asset_type"],
                candidate.get("mapped_answer_asset_id") or "",
                candidate["visual_summary"],
                candidate["image_search_text"],
            ]
        )

    for sheet in [worksheet, summary_sheet, detail_sheet, candidate_sheet]:
        style_header(sheet)
        adjust_width(sheet, max_width=90)

    workbook.save(output_excel_file)
    print(f"UPDATED Excel: {output_excel_file}")


def parse_args() -> argparse.Namespace:
    """解析命令行参数。"""

    parser = argparse.ArgumentParser(
        description="Evaluate query-to-image-description vector search for 04_query_answer+confusion.xlsx."
    )
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL_FILE), help="源 Excel 文件路径。")
    parser.add_argument(
        "--output-excel",
        default=str(DEFAULT_OUTPUT_EXCEL_FILE),
        help="回填后的 Excel 输出路径；传空字符串则不写 Excel。",
    )
    parser.add_argument("--cache-file", default=str(DEFAULT_CACHE_FILE), help="缓存文件路径。")
    parser.add_argument("--dataset", default=str(DEFAULT_DATASET_FILE), help="数据集 JSON 输出路径。")
    parser.add_argument("--result-json", default=str(DEFAULT_RESULT_JSON_FILE), help="结果 JSON 输出路径。")
    parser.add_argument("--result-md", default=str(DEFAULT_RESULT_MD_FILE), help="结果 Markdown 输出路径。")
    parser.add_argument("--top-k", type=int, default=5, help="TopK 命中率。")
    parser.add_argument("--vision-model", default=None, help="视觉模型 endpoint id；不传则从 .env 读取。")
    parser.add_argument("--timeout", type=int, default=180, help="接口超时时间。")
    return parser.parse_args()


def main() -> None:
    """脚本入口。"""

    args = parse_args()
    output_excel_file = Path(args.output_excel) if args.output_excel else None
    result = build_dataset_and_evaluate(
        excel_file=Path(args.excel),
        cache_file=Path(args.cache_file),
        dataset_file=Path(args.dataset),
        result_json_file=Path(args.result_json),
        result_md_file=Path(args.result_md),
        output_excel_file=output_excel_file,
        top_k=args.top_k,
        vision_model=args.vision_model,
        timeout=args.timeout,
    )
    metrics = result["metrics"]
    print("=== Image Semantic Vector Eval Summary ===")
    print(f"total={metrics['total']}")
    print(f"hit_count={metrics['hit_count']}")
    print(f"hit@{args.top_k}={metrics['hit_at_k']:.2%}")
    print(f"hit@1={metrics['hit_at_1']:.2%}")
    print(f"hit@3={metrics['hit_at_3']:.2%}")
    print(f"mrr={metrics['mrr']:.4f}")
    print(f"result_json={args.result_json}")
    print(f"result_md={args.result_md}")
    if output_excel_file is not None:
        print(f"output_excel={output_excel_file}")


if __name__ == "__main__":
    main()
