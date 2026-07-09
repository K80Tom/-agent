"""用 Excel 嵌入图片生成视觉语义描述，并评测 query 到图片描述的向量命中率。

流程：
1. 读取 query_answer 工作表。
2. 取每行 answer_image_file 单元格附近的嵌入图片。
3. 调用火山方舟豆包视觉模型生成图片语义描述。
4. 将图片语义描述向量化，作为候选召回块。
5. 将 query 字段向量化，和全部图片语义描述做余弦相似度排序。
6. 将 Top5 召回排名、分数、召回块写回 Excel。

这个脚本不写数据库，也不写 Milvus。
"""

from __future__ import annotations

import argparse
import base64
import json
import math
import re
import shutil
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import httpx
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.services.vector.doubao_embedding_service import DoubaoEmbeddingService


DOCS_DIR = Path("docs") / "rag提升文档"
DEFAULT_EXCEL_FILE = Path.home() / "Desktop" / "01_query_answer_20_新(1).xlsx"
DEFAULT_CACHE_FILE = DOCS_DIR / "query_answer_20_image_semantic_cache.json"
DEFAULT_DATASET_FILE = DOCS_DIR / "query_answer_20_image_semantic_dataset.json"
DEFAULT_RESULT_JSON_FILE = DOCS_DIR / "query_answer_20_image_semantic_eval_result.json"
DEFAULT_RESULT_MD_FILE = DOCS_DIR / "query_answer_20_image_semantic_eval_result.md"


@dataclass(slots=True)
class ImageCandidate:
    """一个图片语义描述候选块。"""

    query_id: str
    answer_asset_id: str
    answer_asset: str
    visual_summary: str
    search_text: str
    vector: list[float]


def clean_text(value: Any) -> str:
    """把 Excel 单元格或模型输出转成单行稳定文本。"""

    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\r", " ").replace("\n", " ")).strip()


def read_env(path: Path) -> dict[str, str]:
    """读取 .env，保留用户自定义的 endpoint 变量名。"""

    data: dict[str, str] = {}
    if not path.exists():
        return data

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        data[key.strip()] = value.strip().strip('"').strip("'")
    return data


def load_cache(path: Path) -> dict[str, Any]:
    """读取视觉理解和 embedding 缓存。"""

    if not path.exists():
        return {"vision": {}, "embedding": {}}
    data = json.loads(path.read_text(encoding="utf-8"))
    data.setdefault("vision", {})
    data.setdefault("embedding", {})
    return data


def save_cache(path: Path, cache: dict[str, Any]) -> None:
    """保存缓存。"""

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(cache, ensure_ascii=False) + "\n", encoding="utf-8")


def get_sheet_rows(workbook: openpyxl.Workbook) -> list[dict[str, Any]]:
    """读取 query_answer 表数据，并保留 Excel 行号。"""

    worksheet = workbook["query_answer"]
    headers = [clean_text(cell.value) for cell in worksheet[1]]
    header_index = {header: index for index, header in enumerate(headers)}
    required = {
        "query_id",
        "query",
        "answer_asset_id",
        "answer_asset",
        "drama",
        "asset_type",
        "set_id",
        "set_name",
        "answer_image_file",
    }
    missing = sorted(required.difference(header_index))
    if missing:
        raise ValueError(f"Excel 缺少必要列：{', '.join(missing)}")

    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if not any(values):
            continue
        row = {
            field: clean_text(values[header_index[field]])
            for field in required
        }
        row["excel_row"] = row_number
        rows.append(row)
    return rows


def image_bytes_by_row(workbook: openpyxl.Workbook) -> dict[int, bytes]:
    """按 Excel 行号提取嵌入图片二进制。"""

    worksheet = workbook["query_answer"]
    result: dict[int, bytes] = {}
    for image in getattr(worksheet, "_images", []):
        row_number = image.anchor._from.row + 1
        result[row_number] = image._data()
    return result


def vision_prompt(row: dict[str, Any]) -> str:
    """构造图片理解提示词。"""

    return f"""
你是短剧资产库的图片标注员。请只根据图片内容生成中文语义描述，不要编造图片中看不到的信息。

已知元数据只用于辅助判断，不要把答案 ID 当成视觉特征：
- 标准资产名：{row["answer_asset"]}
- 剧名：{row["drama"]}
- 资产类型：{row["asset_type"]}
- 检索集合：{row["set_name"]}

请输出严格 JSON，不要 Markdown，不要解释：
{{
  "visual_summary": "一句话描述图片主体、风格和关键视觉特征",
  "subject": "主体是什么，例如都市男主、古风仙尊、医院病房、诡秘女护士",
  "asset_type": "角色/角色变体/场景/道具/未知",
  "gender": "男/女/未知/不适用",
  "age_range": "年龄段或不适用",
  "hair": "发型、发色、头部特征",
  "clothing": "服装、鞋帽、配饰",
  "colors": ["主色1", "主色2"],
  "scene_space": "场景空间或背景",
  "style": ["都市短剧", "仙侠", "诡秘", "仿真人", "古风", "赛博"],
  "mood": "气质氛围",
  "props": ["道具"],
  "search_tags": ["适合检索的中文标签"],
  "search_text": "把以上信息合并成一段适合向量检索的中文描述"
}}
""".strip()


def parse_json_object(text: str) -> dict[str, Any]:
    """尽量从模型输出中解析 JSON 对象。"""

    content = text.strip()
    if content.startswith("```"):
        content = re.sub(r"^```(?:json)?", "", content).strip()
        content = re.sub(r"```$", "", content).strip()

    try:
        parsed = json.loads(content)
        if isinstance(parsed, dict):
            return parsed
    except json.JSONDecodeError:
        pass

    match = re.search(r"\{.*\}", content, flags=re.S)
    if match:
        try:
            parsed = json.loads(match.group(0))
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            pass

    return {
        "visual_summary": clean_text(text),
        "search_text": clean_text(text),
        "search_tags": [],
    }


def describe_image(
    *,
    row: dict[str, Any],
    image_bytes: bytes,
    cache: dict[str, Any],
    model: str,
    api_key: str,
    base_url: str,
    timeout: int,
) -> dict[str, Any]:
    """调用视觉模型生成图片描述，并缓存。"""

    cache_key = f"{row['query_id']}:{row['answer_asset_id']}:{len(image_bytes)}"
    if cache_key in cache["vision"]:
        return cache["vision"][cache_key]

    image_b64 = base64.b64encode(image_bytes).decode("ascii")
    payload = {
        "model": model,
        "messages": [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": vision_prompt(row)},
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:image/png;base64,{image_b64}"},
                    },
                ],
            }
        ],
        "temperature": 0,
    }
    response = httpx.post(
        f"{base_url.rstrip('/')}/chat/completions",
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=timeout,
    )
    if response.status_code >= 400:
        raise RuntimeError(
            f"Vision model failed for {row['query_id']}: "
            f"status={response.status_code}, body={response.text}"
        )

    data = response.json()
    content = data["choices"][0]["message"]["content"]
    parsed = parse_json_object(content)
    parsed["raw_content"] = content
    parsed["model"] = data.get("model", model)
    cache["vision"][cache_key] = parsed
    return parsed


def embedding_for_text(
    *,
    service: DoubaoEmbeddingService,
    cache: dict[str, Any],
    prefix: str,
    text: str,
) -> list[float]:
    """向量化文本，并缓存。"""

    key = f"{prefix}:{text}"
    if key not in cache["embedding"]:
        cache["embedding"][key] = service.embed_text(text)
    return [float(item) for item in cache["embedding"][key]]


def cosine_similarity(left: list[float], right: list[float]) -> float:
    """计算余弦相似度。"""

    dot = sum(a * b for a, b in zip(left, right))
    left_norm = math.sqrt(sum(a * a for a in left))
    right_norm = math.sqrt(sum(b * b for b in right))
    if left_norm == 0 or right_norm == 0:
        return 0.0
    return dot / (left_norm * right_norm)


def build_dataset_and_evaluate(
    *,
    excel_file: Path,
    cache_file: Path,
    dataset_file: Path,
    result_json_file: Path,
    result_md_file: Path,
    top_k: int,
    vision_model: str | None,
    timeout: int,
) -> dict[str, Any]:
    """执行完整图片语义评测。"""

    env = read_env(PROJECT_ROOT / ".env")
    api_key = env.get("ARK_API_KEY")
    if not api_key:
        raise ValueError("Missing ARK_API_KEY in .env")

    model = (
        vision_model
        or env.get("DOUBAO_LLM_MODEL_2_0_LITE")
        or env.get("DOUBAO_LLM_MODE_2.0lite")
        or env.get("DOUBAO_VISION_MODEL")
        or env.get("DOUBAO_LLM_MODEL")
    )
    if not model:
        raise ValueError("Missing vision model endpoint in .env")

    base_url = env.get("ARK_BASE_URL", "https://ark.cn-beijing.volces.com/api/v3")
    workbook = openpyxl.load_workbook(excel_file)
    rows = get_sheet_rows(workbook)
    row_images = image_bytes_by_row(workbook)
    cache = load_cache(cache_file)
    embedding_service = DoubaoEmbeddingService()

    cases: list[dict[str, Any]] = []
    candidates: list[ImageCandidate] = []

    for row in rows:
        image_bytes = row_images.get(row["excel_row"])
        if not image_bytes:
            raise ValueError(f"{row['query_id']} 第 {row['excel_row']} 行没有嵌入图片")

        visual = describe_image(
            row=row,
            image_bytes=image_bytes,
            cache=cache,
            model=model,
            api_key=api_key,
            base_url=base_url,
            timeout=timeout,
        )
        visual_summary = clean_text(visual.get("visual_summary"))
        search_text = clean_text(visual.get("search_text")) or visual_summary
        if not search_text:
            search_text = clean_text(visual.get("raw_content"))

        vector = embedding_for_text(
            service=embedding_service,
            cache=cache,
            prefix="image_search_text",
            text=search_text,
        )
        case = {
            **row,
            "visual": visual,
            "visual_summary": visual_summary,
            "image_search_text": search_text,
        }
        cases.append(case)
        candidates.append(
            ImageCandidate(
                query_id=row["query_id"],
                answer_asset_id=row["answer_asset_id"],
                answer_asset=row["answer_asset"],
                visual_summary=visual_summary,
                search_text=search_text,
                vector=vector,
            )
        )

    eval_items: list[dict[str, Any]] = []
    for case in cases:
        query_vector = embedding_for_text(
            service=embedding_service,
            cache=cache,
            prefix="query",
            text=case["query"],
        )
        ranked = sorted(
            (
                {
                    "query_id": candidate.query_id,
                    "answer_asset_id": candidate.answer_asset_id,
                    "answer_asset": candidate.answer_asset,
                    "visual_summary": candidate.visual_summary,
                    "image_search_text": candidate.search_text,
                    "score": cosine_similarity(query_vector, candidate.vector),
                }
                for candidate in candidates
            ),
            key=lambda item: item["score"],
            reverse=True,
        )

        expected_id = case["answer_asset_id"]
        hit_rank = None
        hit_score = None
        for rank, candidate in enumerate(ranked, start=1):
            if candidate["answer_asset_id"] == expected_id:
                hit_rank = rank
                hit_score = candidate["score"]
                break

        eval_items.append(
            {
                "query_id": case["query_id"],
                "query_text": case["query"],
                "expected_answer_asset_id": expected_id,
                "expected_answer_asset": case["answer_asset"],
                "visual_summary": case["visual_summary"],
                "image_search_text": case["image_search_text"],
                "hit_rank": hit_rank,
                "hit_score": hit_score,
                "hit_at_top_k": hit_rank is not None and hit_rank <= top_k,
                "top_k": ranked[:top_k],
            }
        )

    total = len(eval_items)
    hit_count = sum(1 for item in eval_items if item["hit_at_top_k"])
    hit_at_1 = sum(1 for item in eval_items if item["hit_rank"] == 1)
    hit_at_3 = sum(
        1 for item in eval_items if item["hit_rank"] is not None and item["hit_rank"] <= 3
    )
    mrr = sum(
        0.0 if item["hit_rank"] is None else 1.0 / item["hit_rank"]
        for item in eval_items
    ) / total

    dataset = {
        "dataset_name": excel_file.stem,
        "source_excel": str(excel_file),
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "vision_model": model,
        "vectorization_rule": {
            "query_text": "使用 query 字段原文向量化。",
            "candidate_text": "使用图片视觉模型生成的 image_search_text 向量化。",
            "hit_rule": "按 answer_asset_id 判断正确图片语义块是否进入 TopK。",
        },
        "cases": cases,
    }
    result = {
        "dataset_name": excel_file.stem,
        "evaluated_at": datetime.now().isoformat(timespec="seconds"),
        "vision_model": model,
        "top_k": top_k,
        "metrics": {
            "total": total,
            "hit_count": hit_count,
            "hit_at_k": hit_count / total if total else 0.0,
            "hit_at_1": hit_at_1 / total if total else 0.0,
            "hit_at_3": hit_at_3 / total if total else 0.0,
            "mrr": mrr,
        },
        "items": eval_items,
    }

    dataset_file.parent.mkdir(parents=True, exist_ok=True)
    dataset_file.write_text(
        json.dumps(dataset, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    result_json_file.parent.mkdir(parents=True, exist_ok=True)
    result_json_file.write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    write_markdown_result(result=result, path=result_md_file)
    save_cache(cache_file, cache)
    backfill_excel(excel_file=excel_file, result=result)
    return result


def write_markdown_result(*, result: dict[str, Any], path: Path) -> None:
    """写 Markdown 结果报告。"""

    metrics = result["metrics"]
    lines = [
        "# 图片语义描述向量检索评测结果",
        "",
        f"- 数据集：`{result['dataset_name']}`",
        f"- 视觉模型：`{result['vision_model']}`",
        f"- TopK：`{result['top_k']}`",
        f"- 评测时间：{result['evaluated_at']}",
        "",
        "## 指标",
        "",
        "| 指标 | 数值 |",
        "|---|---:|",
        f"| total | {metrics['total']} |",
        f"| hit_count | {metrics['hit_count']} |",
        f"| Hit@{result['top_k']} | {metrics['hit_at_k']:.2%} |",
        f"| Hit@1 | {metrics['hit_at_1']:.2%} |",
        f"| Hit@3 | {metrics['hit_at_3']:.2%} |",
        f"| MRR | {metrics['mrr']:.4f} |",
        "",
        "## 明细",
        "",
        "| Query ID | 标准答案 | 命中排名 | 命中分数 | Top1 | Top1 分数 | Top5 |",
        "|---|---|---:|---:|---|---:|---|",
    ]
    for item in result["items"]:
        top_items = item["top_k"]
        top1 = top_items[0] if top_items else {}
        top5 = "；".join(
            f"{rank}.{candidate['answer_asset']}({candidate['score']:.4f})"
            for rank, candidate in enumerate(top_items, start=1)
        )
        lines.append(
            "| {query_id} | {expected} | {hit_rank} | {hit_score} | {top1} | {top1_score} | {top5} |".format(
                query_id=item["query_id"],
                expected=item["expected_answer_asset"],
                hit_rank="" if item["hit_rank"] is None else item["hit_rank"],
                hit_score="" if item["hit_score"] is None else f"{item['hit_score']:.6f}",
                top1=top1.get("answer_asset", ""),
                top1_score="" if "score" not in top1 else f"{top1['score']:.6f}",
                top5=top5,
            )
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def ensure_headers(worksheet: Any, headers: list[str]) -> dict[str, int]:
    """确保 query_answer 表有回填列。"""

    existing = {clean_text(cell.value): index for index, cell in enumerate(worksheet[1], start=1)}
    for header in headers:
        if header not in existing:
            column = worksheet.max_column + 1
            worksheet.cell(row=1, column=column, value=header)
            existing[header] = column
    return existing


def style_header(worksheet: Any) -> None:
    """设置表头样式。"""

    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def adjust_width(worksheet: Any, max_width: int = 60) -> None:
    """自动列宽和自动换行。"""

    for column_cells in worksheet.columns:
        letter = column_cells[0].column_letter
        width = max(len(clean_text(cell.value)) for cell in column_cells) + 2
        worksheet.column_dimensions[letter].width = min(max(width, 10), max_width)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def backfill_excel(*, excel_file: Path, result: dict[str, Any]) -> None:
    """把图片语义评测结果回填到原 Excel。"""

    backup = excel_file.with_name(
        f"{excel_file.stem}.backup_image_semantic_{datetime.now().strftime('%Y%m%d_%H%M%S')}{excel_file.suffix}"
    )
    shutil.copy2(excel_file, backup)

    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook["query_answer"]
    result_by_query_id = {item["query_id"]: item for item in result["items"]}

    headers = [
        "图片语义描述",
        "图片向量文本",
        "图片语义命中排名",
        "图片语义命中分数",
        "图片语义是否Top5命中",
        "图片语义Top1召回块",
        "图片语义Top1分数",
        "图片语义Top5召回块",
        "图片语义Top5分数",
        "图片语义Top5明细",
    ]
    columns = ensure_headers(worksheet, headers)
    source_columns = {
        clean_text(cell.value): index for index, cell in enumerate(worksheet[1], start=1)
    }
    query_id_column = source_columns["query_id"]

    for row_number in range(2, worksheet.max_row + 1):
        query_id = clean_text(worksheet.cell(row=row_number, column=query_id_column).value)
        item = result_by_query_id.get(query_id)
        if not item:
            continue

        top_k = item["top_k"]
        top1 = top_k[0] if top_k else {}
        top5_blocks = "；".join(
            f"{rank}.{candidate['answer_asset']}"
            for rank, candidate in enumerate(top_k, start=1)
        )
        top5_scores = "；".join(
            f"{rank}.{candidate['score']:.6f}"
            for rank, candidate in enumerate(top_k, start=1)
        )
        top5_detail = "；".join(
            f"{rank}.{candidate['answer_asset']}({candidate['answer_asset_id']}, {candidate['score']:.6f})"
            for rank, candidate in enumerate(top_k, start=1)
        )

        worksheet.cell(row=row_number, column=columns["图片语义描述"], value=item["visual_summary"])
        worksheet.cell(row=row_number, column=columns["图片向量文本"], value=item["image_search_text"])
        worksheet.cell(row=row_number, column=columns["图片语义命中排名"], value=item["hit_rank"])
        worksheet.cell(row=row_number, column=columns["图片语义命中分数"], value=item["hit_score"])
        worksheet.cell(
            row=row_number,
            column=columns["图片语义是否Top5命中"],
            value="是" if item["hit_at_top_k"] else "否",
        )
        worksheet.cell(
            row=row_number,
            column=columns["图片语义Top1召回块"],
            value=top1.get("answer_asset", ""),
        )
        worksheet.cell(
            row=row_number,
            column=columns["图片语义Top1分数"],
            value=top1.get("score"),
        )
        worksheet.cell(row=row_number, column=columns["图片语义Top5召回块"], value=top5_blocks)
        worksheet.cell(row=row_number, column=columns["图片语义Top5分数"], value=top5_scores)
        worksheet.cell(row=row_number, column=columns["图片语义Top5明细"], value=top5_detail)

    for sheet_name in ["image_semantic_summary", "image_semantic_top5"]:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    summary_sheet = workbook.create_sheet("image_semantic_summary")
    summary_sheet.append(["指标", "数值"])
    metrics = result["metrics"]
    for metric, value in [
        ("total", metrics["total"]),
        ("hit_count", metrics["hit_count"]),
        (f"Hit@{result['top_k']}", metrics["hit_at_k"]),
        ("Hit@1", metrics["hit_at_1"]),
        ("Hit@3", metrics["hit_at_3"]),
        ("MRR", metrics["mrr"]),
        ("vision_model", result["vision_model"]),
    ]:
        summary_sheet.append([metric, value])

    detail_sheet = workbook.create_sheet("image_semantic_top5")
    detail_sheet.append(
        [
            "query_id",
            "query_text",
            "expected_answer_asset_id",
            "expected_answer_asset",
            "hit_rank",
            "hit_score",
            "hit_at_top5",
            "recall_rank",
            "recall_answer_asset_id",
            "recall_answer_asset",
            "recall_score",
            "is_expected",
            "recall_visual_summary",
            "recall_image_search_text",
        ]
    )
    for item in result["items"]:
        expected_id = item["expected_answer_asset_id"]
        for rank, candidate in enumerate(item["top_k"], start=1):
            candidate_id = candidate["answer_asset_id"]
            detail_sheet.append(
                [
                    item["query_id"],
                    item["query_text"],
                    expected_id,
                    item["expected_answer_asset"],
                    item["hit_rank"],
                    item["hit_score"],
                    "是" if item["hit_at_top_k"] else "否",
                    rank,
                    candidate_id,
                    candidate["answer_asset"],
                    candidate["score"],
                    "是" if candidate_id == expected_id else "否",
                    candidate["visual_summary"],
                    candidate["image_search_text"],
                ]
            )

    for sheet in [worksheet, summary_sheet, detail_sheet]:
        style_header(sheet)
        adjust_width(sheet, max_width=80)

    workbook.save(excel_file)
    print(f"UPDATED Excel: {excel_file}")
    print(f"BACKUP Excel: {backup}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Evaluate query-to-image-visual-description vector search for an Excel dataset."
    )
    parser.add_argument(
        "--excel",
        default=str(DEFAULT_EXCEL_FILE),
        help="带嵌入图片的 Excel 文件路径。",
    )
    parser.add_argument(
        "--cache-file",
        default=str(DEFAULT_CACHE_FILE),
        help="视觉理解和 embedding 缓存文件。",
    )
    parser.add_argument(
        "--dataset",
        default=str(DEFAULT_DATASET_FILE),
        help="图片语义数据集 JSON 输出路径。",
    )
    parser.add_argument(
        "--result-json",
        default=str(DEFAULT_RESULT_JSON_FILE),
        help="评测结果 JSON 输出路径。",
    )
    parser.add_argument(
        "--result-md",
        default=str(DEFAULT_RESULT_MD_FILE),
        help="评测结果 Markdown 输出路径。",
    )
    parser.add_argument("--top-k", type=int, default=5, help="TopK 命中率。")
    parser.add_argument(
        "--vision-model",
        default=None,
        help="视觉模型 endpoint id；不传则从 .env 读取。",
    )
    parser.add_argument("--timeout", type=int, default=180, help="接口超时时间。")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    result = build_dataset_and_evaluate(
        excel_file=Path(args.excel),
        cache_file=Path(args.cache_file),
        dataset_file=Path(args.dataset),
        result_json_file=Path(args.result_json),
        result_md_file=Path(args.result_md),
        top_k=args.top_k,
        vision_model=args.vision_model,
        timeout=args.timeout,
    )
    metrics = result["metrics"]
    print("=== Image Semantic Vector Eval Summary ===")
    print(f"total={metrics['total']}")
    print(f"hit_count={metrics['hit_count']}")
    print(f"hit@{args.top_k}={metrics['hit_at_k']:.2%}")
    print(f"hit@1={metrics['hit_at_1']:.2%}")
    print(f"hit@3={metrics['hit_at_3']:.2%}")
    print(f"mrr={metrics['mrr']:.4f}")
    print(f"result_json={args.result_json}")
    print(f"result_md={args.result_md}")


if __name__ == "__main__":
    main()
