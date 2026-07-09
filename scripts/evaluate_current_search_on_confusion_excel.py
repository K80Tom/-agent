"""用 04_query_answer+confusion.xlsx 评测当前资产检索服务。

这个脚本评测的是当前业务检索链路：
QueryUnderstanding -> 多 query 向量召回 -> SQL 辅助召回 -> RRF 融合排序。

评测难点：
04 Excel 里的标准答案是 ANS001-ANS020，但当前业务检索返回的是数据库
source_table/source_id。因此脚本会先用标准答案图片和资产库图片做感知哈希匹配，
把能匹配的 ANS 映射到真实数据库资产 ID，再计算 Hit@K。

映射不到当前资产库的样本不会计入正式指标，但会输出 Top5 供人工复核。
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import sys
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import httpx
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.db.session import create_session
from app.models.asset_entity_model import AssetEntity
from app.models.asset_variant_model import AssetVariant
from app.services.vector.asset_vector_search_service import AssetVectorSearchService


DOCS_DIR = Path("docs") / "rag提升文档"
DEFAULT_EXCEL_FILE = Path.home() / "Desktop" / "04_query_answer+confusion.xlsx"
DEFAULT_RESULT_JSON = DOCS_DIR / "04_query_answer_confusion_current_search_eval_result.json"
DEFAULT_RESULT_MD = DOCS_DIR / "04_query_answer_confusion_current_search_eval_result.md"
DEFAULT_OUTPUT_EXCEL = Path.home() / "Desktop" / "04_query_answer+confusion_current_search_eval_result.xlsx"
DEFAULT_HASH_CACHE = DOCS_DIR / "04_query_answer_confusion_asset_image_hash_cache.json"


@dataclass(slots=True)
class QueryCase:
    """一条 Excel 评测 query。"""

    excel_row: int
    query_id: str
    query: str
    answer_asset_id: str
    asset_type: str
    image_hash: str
    image_dhash: int


@dataclass(slots=True)
class AssetImage:
    """当前资产库中可用于匹配标准答案图的一张资产图。"""

    source_table: str
    source_id: str
    name: str
    asset_kind: str
    source_file_url: str
    image_dhash: int
    image_sha256: str


def clean_text(value: Any) -> str:
    """把单元格值转为稳定的单行文本。"""

    if value is None:
        return ""
    return str(value).replace("\r", " ").replace("\n", " ").strip()


def dhash_bytes(data: bytes, hash_size: int = 16) -> int:
    """计算图片 dHash，用于压缩/缩放后仍能匹配相似图片。"""

    image = Image.open(BytesIO(data)).convert("L")
    image = image.resize((hash_size + 1, hash_size), Image.Resampling.LANCZOS)
    pixels = list(image.getdata())

    value = 0
    for row in range(hash_size):
        offset = row * (hash_size + 1)
        for col in range(hash_size):
            value = (value << 1) | (
                1 if pixels[offset + col] > pixels[offset + col + 1] else 0
            )
    return value


def hamming_distance(left: int, right: int) -> int:
    """计算两个 dHash 的汉明距离，越小表示图片越像。"""

    return (left ^ right).bit_count()


def load_cases(excel_file: Path) -> list[QueryCase]:
    """读取 Excel 前 20 条 query，并提取对应标准答案图片 hash。"""

    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook["test"] if "test" in workbook.sheetnames else workbook.active

    image_bytes_by_row: dict[int, bytes] = {}
    for image in getattr(worksheet, "_images", []):
        row_number = image.anchor._from.row + 1
        if 2 <= row_number <= 21:
            image_bytes_by_row[row_number] = image._data()

    cases: list[QueryCase] = []
    for row in range(2, 22):
        image_bytes = image_bytes_by_row.get(row)
        if not image_bytes:
            continue
        cases.append(
            QueryCase(
                excel_row=row,
                query_id=clean_text(worksheet.cell(row, 1).value),
                query=clean_text(worksheet.cell(row, 2).value),
                answer_asset_id=clean_text(worksheet.cell(row, 3).value),
                asset_type=clean_text(worksheet.cell(row, 4).value),
                image_hash=hashlib.sha256(image_bytes).hexdigest(),
                image_dhash=dhash_bytes(image_bytes),
            )
        )
    return cases


def load_hash_cache(path: Path) -> dict[str, Any]:
    """读取资产图片 hash 缓存，避免重复下载 TOS 图片。"""

    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def save_hash_cache(path: Path, cache: dict[str, Any]) -> None:
    """保存资产图片 hash 缓存。"""

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(cache, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def collect_asset_image_records(db: Any) -> list[dict[str, str]]:
    """读取当前资产库里带 source_file_url 的实体和变体。"""

    records: list[dict[str, str]] = []
    for entity in db.query(AssetEntity).filter(AssetEntity.source_file_url.isnot(None)).all():
        records.append(
            {
                "source_table": "asset_entities",
                "source_id": str(entity.id),
                "name": entity.name,
                "asset_kind": entity.asset_kind,
                "source_file_url": entity.source_file_url,
            }
        )

    for variant in db.query(AssetVariant).filter(AssetVariant.source_file_url.isnot(None)).all():
        metadata = variant.metadata_ or {}
        parent_name = clean_text(metadata.get("parent_name"))
        display_name = f"{parent_name}/{variant.name}" if parent_name else variant.name
        records.append(
            {
                "source_table": "asset_variants",
                "source_id": str(variant.id),
                "name": display_name,
                "asset_kind": variant.variant_kind,
                "source_file_url": variant.source_file_url,
            }
        )
    return records


def build_asset_images(
    *,
    db: Any,
    cache_file: Path,
    timeout: int,
) -> list[AssetImage]:
    """下载并计算当前资产库图片 dHash。"""

    cache = load_hash_cache(cache_file)
    records = collect_asset_image_records(db)
    client = httpx.Client(timeout=timeout, follow_redirects=True)
    asset_images: list[AssetImage] = []

    for index, record in enumerate(records, start=1):
        cache_key = f"{record['source_table']}:{record['source_id']}:{record['source_file_url']}"
        cached = cache.get(cache_key)
        if cached is None:
            try:
                response = client.get(record["source_file_url"])
                if response.status_code >= 400:
                    continue
                image_bytes = response.content
                cached = {
                    "image_sha256": hashlib.sha256(image_bytes).hexdigest(),
                    "image_dhash": str(dhash_bytes(image_bytes)),
                }
                cache[cache_key] = cached
                if index % 10 == 0:
                    save_hash_cache(cache_file, cache)
            except Exception as exc:
                print(f"[image-hash-skip] {record['source_table']} {record['name']}: {exc}")
                continue

        asset_images.append(
            AssetImage(
                source_table=record["source_table"],
                source_id=record["source_id"],
                name=record["name"],
                asset_kind=record["asset_kind"],
                source_file_url=record["source_file_url"],
                image_sha256=str(cached["image_sha256"]),
                image_dhash=int(cached["image_dhash"]),
            )
        )

    save_hash_cache(cache_file, cache)
    return asset_images


def map_expected_assets(
    *,
    cases: list[QueryCase],
    asset_images: list[AssetImage],
    max_distance: int,
) -> dict[str, list[dict[str, Any]]]:
    """把 Excel 标准答案图片映射到当前资产库 ID。"""

    mappings: dict[str, list[dict[str, Any]]] = {}
    for case in cases:
        ranked = sorted(
            (
                {
                    "source_table": asset.source_table,
                    "source_id": asset.source_id,
                    "name": asset.name,
                    "asset_kind": asset.asset_kind,
                    "image_distance": hamming_distance(case.image_dhash, asset.image_dhash),
                }
                for asset in asset_images
            ),
            key=lambda item: item["image_distance"],
        )
        best_distance = ranked[0]["image_distance"] if ranked else None
        if best_distance is not None and best_distance <= max_distance:
            mappings[case.query_id] = [
                item for item in ranked if item["image_distance"] <= max_distance
            ]
        else:
            mappings[case.query_id] = []
    return mappings


def find_hit_rank(
    *,
    expected_assets: list[dict[str, Any]],
    returned_items: list[dict[str, Any]],
) -> int | None:
    """按 source_table/source_id 判断当前检索是否命中标准答案。"""

    expected_keys = {
        (item["source_table"], item["source_id"])
        for item in expected_assets
    }
    for rank, item in enumerate(returned_items, start=1):
        key = (str(item.get("source_table")), str(item.get("source_id")))
        if key in expected_keys:
            return rank
    return None


def compact_item(item: dict[str, Any]) -> dict[str, Any]:
    """压缩检索结果，避免结果文件过大。"""

    metadata = item.get("metadata") or {}
    debug = metadata.get("search_debug") or {}
    return {
        "score": item.get("score"),
        "source_table": item.get("source_table"),
        "source_id": item.get("source_id"),
        "asset_kind": item.get("asset_kind"),
        "name": item.get("name"),
        "display_name": item.get("display_name"),
        "parent_entity_name": item.get("parent_entity_name"),
        "source_file_url": item.get("source_file_url"),
        "raw_fusion_score": debug.get("raw_fusion_score"),
        "fusion_reasons": debug.get("fusion_reasons"),
        "recall_signals": debug.get("recall_signals"),
    }


def evaluate(
    *,
    excel_file: Path,
    result_json: Path,
    result_md: Path,
    output_excel: Path,
    hash_cache: Path,
    limit: int,
    max_image_distance: int,
    download_timeout: int,
) -> dict[str, Any]:
    """执行完整评测并写出结果。"""

    cases = load_cases(excel_file)
    db = create_session()
    try:
        asset_images = build_asset_images(
            db=db,
            cache_file=hash_cache,
            timeout=download_timeout,
        )
        expected_mapping = map_expected_assets(
            cases=cases,
            asset_images=asset_images,
            max_distance=max_image_distance,
        )

        service = AssetVectorSearchService(db)
        items: list[dict[str, Any]] = []
        for index, case in enumerate(cases, start=1):
            print(f"[search] {index}/{len(cases)} {case.query_id}")
            returned = service.search(query=case.query, limit=limit)
            compact_top = [compact_item(item) for item in returned]
            expected_assets = expected_mapping.get(case.query_id) or []
            hit_rank = find_hit_rank(
                expected_assets=expected_assets,
                returned_items=returned,
            )
            items.append(
                {
                    "query_id": case.query_id,
                    "query": case.query,
                    "answer_asset_id": case.answer_asset_id,
                    "asset_type": case.asset_type,
                    "mapped": bool(expected_assets),
                    "expected_assets": expected_assets,
                    "hit_rank": hit_rank,
                    "hit_at_1": hit_rank == 1,
                    "hit_at_3": hit_rank is not None and hit_rank <= 3,
                    "hit_at_5": hit_rank is not None and hit_rank <= 5,
                    "top_k": compact_top,
                }
            )
    finally:
        db.close()

    mapped_items = [item for item in items if item["mapped"]]
    mapped_total = len(mapped_items)
    metrics = {
        "total": len(items),
        "mapped_total": mapped_total,
        "unmapped_total": len(items) - mapped_total,
        "hit_at_1": (
            sum(1 for item in mapped_items if item["hit_at_1"]) / mapped_total
            if mapped_total
            else 0.0
        ),
        "hit_at_3": (
            sum(1 for item in mapped_items if item["hit_at_3"]) / mapped_total
            if mapped_total
            else 0.0
        ),
        "hit_at_5": (
            sum(1 for item in mapped_items if item["hit_at_5"]) / mapped_total
            if mapped_total
            else 0.0
        ),
        "mrr": (
            sum(0.0 if item["hit_rank"] is None else 1.0 / item["hit_rank"] for item in mapped_items)
            / mapped_total
            if mapped_total
            else 0.0
        ),
    }

    result = {
        "dataset_name": excel_file.stem,
        "evaluated_at": datetime.now().isoformat(timespec="seconds"),
        "method": "current_asset_search_service",
        "limit": limit,
        "max_image_distance": max_image_distance,
        "mapping_rule": "Excel 标准答案图片 dHash 与当前资产库 source_file_url 图片 dHash 距离 <= max_image_distance 时，视为可自动映射标准答案。",
        "metrics": metrics,
        "items": items,
    }

    result_json.parent.mkdir(parents=True, exist_ok=True)
    result_json.write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    write_markdown(result=result, path=result_md)
    write_excel(result=result, path=output_excel)
    return result


def display_asset(asset: dict[str, Any]) -> str:
    """格式化资产显示名。"""

    if not asset:
        return ""
    source_table = asset.get("source_table") or ""
    name = asset.get("name") or asset.get("display_name") or ""
    score = asset.get("score")
    suffix = "" if score is None else f"({float(score):.4f})"
    return f"{source_table}/{name}{suffix}"


def write_markdown(*, result: dict[str, Any], path: Path) -> None:
    """写 Markdown 评测报告。"""

    metrics = result["metrics"]
    lines = [
        "# 当前资产检索方法在 04 Excel 上的评测结果",
        "",
        f"- 数据集：`{result['dataset_name']}`",
        f"- 方法：`{result['method']}`",
        f"- TopK：`{result['limit']}`",
        f"- 评测时间：{result['evaluated_at']}",
        f"- 映射规则：{result['mapping_rule']}",
        "",
        "## 指标",
        "",
        "| 指标 | 数值 |",
        "|---|---:|",
        f"| total | {metrics['total']} |",
        f"| mapped_total | {metrics['mapped_total']} |",
        f"| unmapped_total | {metrics['unmapped_total']} |",
        f"| Hit@1 | {metrics['hit_at_1']:.2%} |",
        f"| Hit@3 | {metrics['hit_at_3']:.2%} |",
        f"| Hit@5 | {metrics['hit_at_5']:.2%} |",
        f"| MRR | {metrics['mrr']:.4f} |",
        "",
        "## 明细",
        "",
        "| Query ID | 标准答案 | 映射到当前资产 | 命中排名 | Top1 | Top5 |",
        "|---|---|---|---:|---|---|",
    ]

    for item in result["items"]:
        expected = "；".join(
            f"{asset['source_table']}/{asset['name']}[dist={asset['image_distance']}]"
            for asset in item["expected_assets"]
        )
        top_items = item["top_k"]
        top1 = display_asset(top_items[0]) if top_items else ""
        top5 = "；".join(
            f"{rank}.{display_asset(asset)}"
            for rank, asset in enumerate(top_items, start=1)
        )
        lines.append(
            f"| {item['query_id']} | {item['answer_asset_id']} | {expected or '未映射'} | "
            f"{'' if item['hit_rank'] is None else item['hit_rank']} | {top1} | {top5} |"
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_excel(*, result: dict[str, Any], path: Path) -> None:
    """写 Excel 结果，方便直接查看。"""

    workbook = openpyxl.Workbook()
    summary = workbook.active
    summary.title = "直接看结果"

    metrics = result["metrics"]
    summary.append(["当前资产检索方法评测", ""])
    for row in [
        ["total", metrics["total"]],
        ["mapped_total", metrics["mapped_total"]],
        ["unmapped_total", metrics["unmapped_total"]],
        ["Hit@1", metrics["hit_at_1"]],
        ["Hit@3", metrics["hit_at_3"]],
        ["Hit@5", metrics["hit_at_5"]],
        ["MRR", metrics["mrr"]],
        ["映射规则", result["mapping_rule"]],
    ]:
        summary.append(row)
    summary.append([])
    summary.append(
        [
            "query_id",
            "query",
            "answer_asset_id",
            "映射到当前资产",
            "命中排名",
            "Top1",
            "Top1分数",
            "Top5",
        ]
    )

    for item in result["items"]:
        expected = "；".join(
            f"{asset['source_table']}/{asset['name']}[dist={asset['image_distance']}]"
            for asset in item["expected_assets"]
        )
        top_items = item["top_k"]
        top1 = top_items[0] if top_items else {}
        top5 = "；".join(
            f"{rank}.{display_asset(asset)}"
            for rank, asset in enumerate(top_items, start=1)
        )
        summary.append(
            [
                item["query_id"],
                item["query"],
                item["answer_asset_id"],
                expected or "未映射",
                item["hit_rank"],
                display_asset(top1),
                top1.get("score"),
                top5,
            ]
        )

    detail = workbook.create_sheet("top5_detail")
    detail.append(
        [
            "query_id",
            "answer_asset_id",
            "mapped",
            "hit_rank",
            "rank",
            "source_table",
            "source_id",
            "name",
            "asset_kind",
            "score",
            "raw_fusion_score",
            "fusion_reasons",
            "recall_signals",
        ]
    )
    for item in result["items"]:
        for rank, asset in enumerate(item["top_k"], start=1):
            detail.append(
                [
                    item["query_id"],
                    item["answer_asset_id"],
                    "是" if item["mapped"] else "否",
                    item["hit_rank"],
                    rank,
                    asset.get("source_table"),
                    asset.get("source_id"),
                    asset.get("name") or asset.get("display_name"),
                    asset.get("asset_kind"),
                    asset.get("score"),
                    asset.get("raw_fusion_score"),
                    json.dumps(asset.get("fusion_reasons") or [], ensure_ascii=False),
                    json.dumps(asset.get("recall_signals") or [], ensure_ascii=False),
                ]
            )

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for sheet in [summary, detail]:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in sheet.columns:
            width = max(len(clean_text(cell.value)) for cell in column_cells) + 2
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 10), 80)
        sheet.freeze_panes = "A2"

    workbook.active = 0
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Evaluate current asset search on 04 Excel.")
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL_FILE))
    parser.add_argument("--result-json", default=str(DEFAULT_RESULT_JSON))
    parser.add_argument("--result-md", default=str(DEFAULT_RESULT_MD))
    parser.add_argument("--output-excel", default=str(DEFAULT_OUTPUT_EXCEL))
    parser.add_argument("--hash-cache", default=str(DEFAULT_HASH_CACHE))
    parser.add_argument("--limit", type=int, default=5)
    parser.add_argument("--max-image-distance", type=int, default=5)
    parser.add_argument("--download-timeout", type=int, default=15)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    result = evaluate(
        excel_file=Path(args.excel),
        result_json=Path(args.result_json),
        result_md=Path(args.result_md),
        output_excel=Path(args.output_excel),
        hash_cache=Path(args.hash_cache),
        limit=args.limit,
        max_image_distance=args.max_image_distance,
        download_timeout=args.download_timeout,
    )
    metrics = result["metrics"]
    print("=== Current Asset Search Eval Summary ===")
    print(f"total={metrics['total']}")
    print(f"mapped_total={metrics['mapped_total']}")
    print(f"unmapped_total={metrics['unmapped_total']}")
    print(f"hit@1={metrics['hit_at_1']:.2%}")
    print(f"hit@3={metrics['hit_at_3']:.2%}")
    print(f"hit@5={metrics['hit_at_5']:.2%}")
    print(f"mrr={metrics['mrr']:.4f}")
    print(f"result_json={args.result_json}")
    print(f"result_md={args.result_md}")
    print(f"output_excel={args.output_excel}")


if __name__ == "__main__":
    main()
